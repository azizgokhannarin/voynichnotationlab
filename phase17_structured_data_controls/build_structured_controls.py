#!/usr/bin/env python3
"""Build the preregistered Phase-17 structured-data control corpora."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import random
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


SEED = 20260826
EXPECTED = {
    "instrument": "f4cb1b8e970f6982cbc1a539292ffb188c7f8aa714ad0a437f100aa455a946a4",
    "voynich_corpus": "5fdf577932f21b6da59b7ae12f5bb5451d9bb5b574d81c1affd8b646364b9997",
    "retail_xlsx": "43465a06f2ccf7c8b5bd2892bc7defb52f97487934fe93b16ae4c3936424676d",
    "mushroom_data": "e65d082030501a3ebcbcd7c9f7c71aa9d28fdfff463bf4cf4716a3fe13ac360e",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_hash(value: object) -> str:
    raw = json.dumps(value, ensure_ascii=False, sort_keys=True,
                     separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def stable_u64(text: str) -> int:
    return int.from_bytes(hashlib.sha256(text.encode("utf-8")).digest()[:8], "big")


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
                    encoding="utf-8")


def verify_file(path: Path, expected: str, label: str) -> None:
    actual = sha256_file(path)
    if actual != expected:
        raise ValueError(f"{label} SHA-256 mismatch: {actual} != {expected}")


def validation_layout(voynich_path: Path) -> tuple[list[dict], str, dict]:
    corpus = json.loads(voynich_path.read_text(encoding="utf-8"))
    expected = corpus.pop("corpus_sha256")
    actual = canonical_hash(corpus)
    corpus["corpus_sha256"] = expected
    if expected != EXPECTED["voynich_corpus"] or actual != expected:
        raise ValueError("Voynich canonical corpus mismatch")
    records = [r for r in corpus["records"] if r["split"] == "validation"]
    records.sort(key=lambda r: (r["document"], r["page"], r["order"], r["line_id"]))
    page_map: dict[tuple[str, str], str] = {}
    page_orders = Counter()
    specs = []
    for index, record in enumerate(records):
        key = (record["document"], record["page"])
        if key not in page_map:
            page_map[key] = f"page-{len(page_map):03d}"
        page = page_map[key]
        order = page_orders[page]
        page_orders[page] += 1
        specs.append({"index": index, "page": page, "order": order,
                      "line_id": f"line-{index:04d}", "length": len(record["tokens"])})
    geometry = [{"page": s["page"], "order": s["order"], "length": s["length"]}
                for s in specs]
    counts = {"pages": len(page_map), "lines": len(specs),
              "tokens": sum(s["length"] for s in specs)}
    if counts != {"pages": 45, "lines": 1024, "tokens": 7596}:
        raise ValueError(f"unexpected validation geometry: {counts}")
    return specs, canonical_hash(geometry), counts


def corpus_payload(name: str, specs: list[dict], token_lines: list[list[str]],
                   source: dict) -> dict:
    if len(specs) != len(token_lines):
        raise ValueError("line count mismatch")
    records = []
    for spec, tokens in zip(specs, token_lines):
        if len(tokens) != spec["length"]:
            raise ValueError(f"line length mismatch at {spec['line_id']}")
        records.append({"document": name, "page": spec["page"],
                        "order": spec["order"], "line_id": spec["line_id"],
                        "split": "validation", "tokens": tokens})
    body = {"schema": "PHASE15_LINE_CORPUS_v2", "name": name,
            "source": source, "records": records}
    body["corpus_sha256"] = canonical_hash(body)
    return body


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def complete_invoice(invoice_no: str, invoice_date: object,
                     item_rows: list[tuple[str, int]]) -> tuple[str, str, list[str]] | None:
    if not invoice_no or invoice_no.upper().startswith("C") or not item_rows:
        return None
    tally = []
    for stock_code, quantity in item_rows:
        tally.extend([f"S:{stock_code}"] * min(quantity, 4))
    if not tally:
        return None
    date = invoice_date.isoformat() if hasattr(invoice_date, "isoformat") else cell_text(invoice_date)
    if not date:
        return None
    return invoice_no, date, tally


def retail_line_inventories(xlsx: Path, specs: list[dict]) -> tuple[list[list[str]], dict]:
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [cell_text(value) for value in next(rows)]
    required = {"InvoiceNo", "StockCode", "Quantity", "InvoiceDate"}
    if not required.issubset(header):
        raise ValueError(f"Online Retail header mismatch: {header}")
    position = {name: header.index(name) for name in required}
    selected: list[list[str]] = []
    selected_ids = []
    scanned_rows = 0
    scanned_invoices = 0
    current_no = None
    current_date = None
    current_items: list[tuple[str, int]] = []

    def consider() -> None:
        nonlocal scanned_invoices
        if current_no is None or len(selected) >= len(specs):
            return
        scanned_invoices += 1
        invoice = complete_invoice(current_no, current_date, current_items)
        if invoice is None:
            return
        invoice_no, date, tally = invoice
        needed = specs[len(selected)]["length"]
        if len(tally) >= needed:
            selected.append(tally[:needed])
            selected_ids.append(f"{invoice_no}\t{date}")

    for row in rows:
        scanned_rows += 1
        invoice_no = cell_text(row[position["InvoiceNo"]])
        if current_no is not None and invoice_no != current_no:
            consider()
            if len(selected) >= len(specs):
                break
            current_items = []
        if invoice_no != current_no:
            current_no = invoice_no
            current_date = row[position["InvoiceDate"]]
        stock = cell_text(row[position["StockCode"]])
        try:
            quantity = int(float(row[position["Quantity"]]))
        except (TypeError, ValueError):
            quantity = 0
        if stock and quantity > 0 and not invoice_no.upper().startswith("C"):
            current_items.append((stock, quantity))
    if len(selected) < len(specs):
        consider()
    workbook.close()
    if len(selected) != len(specs):
        raise ValueError(f"not enough eligible invoices: {len(selected)} / {len(specs)}")
    metadata = {
        "selected_invoices": len(selected),
        "selection_sha256": hashlib.sha256(("\n".join(selected_ids) + "\n").encode()).hexdigest(),
        "scanned_rows": scanned_rows,
        "scanned_invoices": scanned_invoices,
        "quantity_cap": 4,
    }
    return selected, metadata


def deterministic_shuffle(lines: list[list[str]], name: str, specs: list[dict]) -> list[list[str]]:
    output = []
    for spec, line in zip(specs, lines):
        shuffled = line[:]
        rng = random.Random(stable_u64(f"{SEED}:{name}:{spec['line_id']}"))
        rng.shuffle(shuffled)
        output.append(shuffled)
    return output


def mushroom_rows(path: Path) -> list[list[str]]:
    rows = []
    with path.open(encoding="utf-8", newline="") as stream:
        for row in csv.reader(stream):
            values = [value.strip() for value in row]
            if len(values) != 23 or any(value == "" for value in values):
                raise ValueError("Mushroom row/schema mismatch")
            rows.append(values)
    if len(rows) != 8124:
        raise ValueError(f"Mushroom row count mismatch: {len(rows)}")
    return rows


def mushroom_line_inventories(rows: list[list[str]], specs: list[dict],
                              qualified: bool) -> tuple[list[list[str]], dict]:
    cursor = 0
    output = []
    used_rows = []
    for spec in specs:
        needed = spec["length"]
        line = []
        while len(line) < needed:
            if cursor >= len(rows):
                raise ValueError("Mushroom source exhausted")
            row = rows[cursor]
            used_rows.append(cursor)
            cursor += 1
            tokens = ([f"F{i:02d}:{value}" for i, value in enumerate(row)]
                      if qualified else row)
            line.extend(tokens[:needed - len(line)])
        output.append(line)
    metadata = {
        "source_rows_consumed": cursor,
        "row_selection_sha256": hashlib.sha256(
            ("\n".join(map(str, used_rows)) + "\n").encode()).hexdigest(),
        "fields_per_source_row": 23,
        "qualified": qualified,
    }
    return output, metadata


def exact_line_inventory_equal(left: dict, right: dict) -> bool:
    return all(Counter(a["tokens"]) == Counter(b["tokens"])
               for a, b in zip(left["records"], right["records"]))


def build_all(args) -> dict:
    instrument = Path(args.instrument)
    voynich = Path(args.voynich)
    retail = Path(args.retail)
    mushroom = Path(args.mushroom)
    verify_file(instrument, EXPECTED["instrument"], "Phase-16 instrument")
    verify_file(retail, EXPECTED["retail_xlsx"], "Online Retail payload")
    verify_file(mushroom, EXPECTED["mushroom_data"], "Mushroom payload")
    specs, geometry_hash, counts = validation_layout(voynich)

    retail_ordered, retail_meta = retail_line_inventories(retail, specs)
    retail_unordered = deterministic_shuffle(retail_ordered,
                                             "RETAIL_TALLY_UNORDERED_v1", specs)
    mushroom_source = mushroom_rows(mushroom)
    mushroom_raw, mushroom_raw_meta = mushroom_line_inventories(mushroom_source, specs, False)
    mushroom_qualified, mushroom_qualified_meta = mushroom_line_inventories(
        mushroom_source, specs, True)
    mushroom_permuted = deterministic_shuffle(mushroom_raw,
                                              "MUSHROOM_RAW_PERMUTED_v1", specs)

    common = {"layout": "Voynich VALIDATION geometry only",
              "layout_sha256": geometry_hash, "geometry_counts": counts,
              "voynich_token_identities_used": False, "voynich_final_test_used": False,
              "seed": SEED}
    corpora = {
        "RETAIL_TALLY_ORDERED_v1": corpus_payload(
            "RETAIL_TALLY_ORDERED_v1", specs, retail_ordered,
            {**common, "dataset": "UCI Online Retail", "doi": "10.24432/C5BW33",
             "payload_sha256": EXPECTED["retail_xlsx"], "encoding": "quantity tally; grouped",
             **retail_meta}),
        "RETAIL_TALLY_UNORDERED_v1": corpus_payload(
            "RETAIL_TALLY_UNORDERED_v1", specs, retail_unordered,
            {**common, "dataset": "UCI Online Retail", "doi": "10.24432/C5BW33",
             "payload_sha256": EXPECTED["retail_xlsx"],
             "encoding": "same quantity-tally line inventories; deterministic order ablation",
             **retail_meta}),
        "MUSHROOM_RAW_ORDERED_v1": corpus_payload(
            "MUSHROOM_RAW_ORDERED_v1", specs, mushroom_raw,
            {**common, "dataset": "UCI Mushroom", "doi": "10.24432/C5959T",
             "payload_sha256": EXPECTED["mushroom_data"],
             "encoding": "raw categorical codes; fixed column order", **mushroom_raw_meta}),
        "MUSHROOM_RAW_PERMUTED_v1": corpus_payload(
            "MUSHROOM_RAW_PERMUTED_v1", specs, mushroom_permuted,
            {**common, "dataset": "UCI Mushroom", "doi": "10.24432/C5959T",
             "payload_sha256": EXPECTED["mushroom_data"],
             "encoding": "same raw-code line inventories; deterministic order ablation",
             **mushroom_raw_meta}),
        "MUSHROOM_QUALIFIED_ORDERED_v1": corpus_payload(
            "MUSHROOM_QUALIFIED_ORDERED_v1", specs, mushroom_qualified,
            {**common, "dataset": "UCI Mushroom", "doi": "10.24432/C5959T",
             "payload_sha256": EXPECTED["mushroom_data"],
             "encoding": "column-qualified categorical codes; fixed column order",
             **mushroom_qualified_meta}),
    }
    if not exact_line_inventory_equal(corpora["RETAIL_TALLY_ORDERED_v1"],
                                      corpora["RETAIL_TALLY_UNORDERED_v1"]):
        raise AssertionError("retail paired-inventory gate failed")
    if not exact_line_inventory_equal(corpora["MUSHROOM_RAW_ORDERED_v1"],
                                      corpora["MUSHROOM_RAW_PERMUTED_v1"]):
        raise AssertionError("mushroom paired-inventory gate failed")

    out_dir = Path(args.out_dir)
    paths = {}
    for name, corpus in corpora.items():
        path = out_dir / f"{name}.json"
        write_json(path, corpus)
        paths[name] = str(path)
    summary = {
        "schema": "PHASE17_STRUCTURED_CONTROL_BUILD_v1",
        "instrument_sha256": EXPECTED["instrument"],
        "layout_sha256": geometry_hash,
        "geometry_counts": counts,
        "source_sha256": {"retail": EXPECTED["retail_xlsx"],
                          "mushroom": EXPECTED["mushroom_data"]},
        "corpus_sha256": {name: corpus["corpus_sha256"] for name, corpus in corpora.items()},
        "paths": paths,
        "paired_inventory_gates": {"retail": True, "mushroom_raw": True},
        "retail_selection": retail_meta,
        "mushroom_raw_selection": mushroom_raw_meta,
        "mushroom_qualified_selection": mushroom_qualified_meta,
        "voynich_final_test_used": False,
    }
    summary["build_sha256"] = canonical_hash(summary)
    write_json(out_dir / "BUILD_SUMMARY_v1.json", summary)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--instrument", required=True)
    parser.add_argument("--voynich", required=True)
    parser.add_argument("--retail", required=True)
    parser.add_argument("--mushroom", required=True)
    parser.add_argument("--out-dir", required=True)
    args = parser.parse_args()
    result = build_all(args)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
